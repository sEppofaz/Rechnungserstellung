#!/usr/bin/env python3
import base64
import difflib
import hashlib
import hmac
import json
import os
import re
import shutil
import subprocess
import tempfile
import threading
import time
import urllib.parse
import urllib.request
import uuid as uuid_module
from datetime import datetime
from functools import wraps
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    _OPENPYXL_SUPPORTED = True
except ImportError:
    _OPENPYXL_SUPPORTED = False

try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
    _HEIC_SUPPORTED = True
except ImportError:
    _HEIC_SUPPORTED = False

import anthropic
import dropbox
from docxtpl import DocxTemplate
from dropbox.files import DeletedMetadata, FolderMetadata
from flask import Flask, Response, abort, request, send_file
from PIL import Image

app = Flask(__name__)

# ── Konfiguration ─────────────────────────────────────────────────────────────

DROPBOX_INVOICE_REFRESH_TOKEN = os.environ["DROPBOX_INVOICE_REFRESH_TOKEN"]
DROPBOX_INVOICE_APP_KEY       = os.environ["DROPBOX_INVOICE_APP_KEY"]
DROPBOX_INVOICE_APP_SECRET    = os.environ["DROPBOX_INVOICE_APP_SECRET"]
CLAUDE_API_KEY                = os.environ["CLAUDE_API_KEY"]
INVOICE_MODEL                 = os.environ.get("CLAUDE_INVOICE_MODEL", "claude-sonnet-4-6")
KARGL_APP_TOKEN               = os.environ.get("KARGL_APP_TOKEN", "")

INVOICE_INPUT_FOLDER  = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Input"
INVOICE_OUTPUT_FOLDER = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Entwurf"
INVOICE_DONE_FOLDER   = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Erledigt"
INVOICE_ERROR_FOLDER  = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/Rechnungen_Fehler"
INVOICE_ADDRESS_FILE  = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/_Adressen.xlsx"
INVOICE_REGISTER_FILE = "/_Austauschordner-Sandra-sEpp/Kargl-Rechnung/_Rechnungsregister.xlsx"

INVOICE_CURSOR_FILE = "/opt/kargl-invoice/invoice_cursor.txt"
INVOICE_TEMPLATE    = "/opt/kargl-invoice/template.docx"
SESSIONS_DIR        = Path("/opt/kargl-invoice/sessions")
KARGL_ICONS_DIR     = Path("/opt/kargl-invoice/icons")
KARGL_HTML_PATH     = Path(__file__).parent / "kargl_app.html"

MEDIA_TYPES = {
    ".pdf":  "application/pdf",
    ".jpg":  "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png":  "image/png",
    ".tiff": "image/tiff",
    ".tif":  "image/tiff",
    ".webp": "image/webp",
    ".bmp":  "image/bmp",
}

_ADDR_HEADERS = ["Name", "Straße", "PLZ", "Ort", "Straße validiert", "PLZ+Ort validiert", "Hinzugefügt"]
_REG_HEADERS  = ["Rechnungsnummer", "Datum", "Anrede", "Nachname", "Vorname",
                 "Produkt", "Netto (€)", "MwSt (€)", "Brutto (€)"]


def log(msg: str) -> None:
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}", flush=True)


# ── Auth ──────────────────────────────────────────────────────────────────────

def require_token(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        if not KARGL_APP_TOKEN or not auth.startswith("Bearer ") or auth[7:] != KARGL_APP_TOKEN:
            return {"error": "Unauthorized"}, 401
        return f(*args, **kwargs)
    return decorated


def require_token_or_param(f):
    """Wie require_token, akzeptiert aber auch ?token= Query-Parameter (für iframe/neuen Tab)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if KARGL_APP_TOKEN:
            auth      = request.headers.get("Authorization", "")
            header_ok = auth.startswith("Bearer ") and auth[7:] == KARGL_APP_TOKEN
            param_ok  = request.args.get("token") == KARGL_APP_TOKEN
            if not header_ok and not param_ok:
                return {"error": "Unauthorized"}, 401
        return f(*args, **kwargs)
    return decorated


# ── ODT-Konvertierung ─────────────────────────────────────────────────────────

def convert_to_odt(docx_path: str) -> tuple[str, str]:
    """Konvertiert .docx → .odt via LibreOffice headless. Gibt (pfad, extension) zurück.
    Fällt auf .docx zurück wenn LibreOffice fehlt."""
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo:
        log("⚠️  LibreOffice nicht gefunden – Fallback .docx")
        return docx_path, ".docx"
    tmp_dir = tempfile.mkdtemp()
    try:
        subprocess.run(
            [lo, "--headless", "--convert-to", "odt", "--outdir", tmp_dir, docx_path],
            check=True, timeout=30, capture_output=True,
        )
        odt_src = Path(tmp_dir) / (Path(docx_path).stem + ".odt")
        if not odt_src.exists():
            log("⚠️  LibreOffice erzeugte keine ODT-Datei – Fallback .docx")
            return docx_path, ".docx"
        final = tempfile.mktemp(suffix=".odt")
        shutil.move(str(odt_src), final)
        return final, ".odt"
    except Exception as e:
        log(f"⚠️  LibreOffice Fehler: {e} – Fallback .docx")
        return docx_path, ".docx"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ── PDF-Konvertierung ─────────────────────────────────────────────────────────

def convert_to_pdf(src_path: str) -> str | None:
    """Konvertiert DOCX/ODT → PDF via LibreOffice. Gibt Pfad zurück oder None bei Fehler."""
    lo = shutil.which("libreoffice") or shutil.which("soffice")
    if not lo:
        return None
    tmp_dir = tempfile.mkdtemp()
    try:
        subprocess.run(
            [lo, "--headless", "--convert-to", "pdf", "--outdir", tmp_dir, src_path],
            check=True, timeout=60, capture_output=True,
        )
        pdf_src = Path(tmp_dir) / (Path(src_path).stem + ".pdf")
        if not pdf_src.exists():
            return None
        final = tempfile.mktemp(suffix=".pdf")
        shutil.move(str(pdf_src), final)
        return final
    except Exception as e:
        log(f"⚠️  PDF-Konvertierung: {e}")
        return None
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ── Session-Management (App-OCR) ──────────────────────────────────────────────

def _cleanup_old_sessions() -> None:
    if not SESSIONS_DIR.exists():
        return
    cutoff = datetime.now().timestamp() - 24 * 3600
    for f in SESSIONS_DIR.iterdir():
        try:
            if f.stat().st_mtime < cutoff:
                f.unlink()
        except Exception:
            pass


# ── App-Icons ─────────────────────────────────────────────────────────────────

def _ensure_kargl_icons() -> None:
    KARGL_ICONS_DIR.mkdir(exist_ok=True)
    version_file = KARGL_ICONS_DIR / ".version"
    current_ok = (
        version_file.exists()
        and version_file.read_text().strip() == _ICON_VERSION
        and all((KARGL_ICONS_DIR / f"icon-{s}.png").exists() for s in (192, 512, 180))
    )
    if current_ok:
        return
    for size in (192, 512, 180):
        p = KARGL_ICONS_DIR / f"icon-{size}.png"
        p.unlink(missing_ok=True)
        _generate_kargl_icon(size, p)
    version_file.write_text(_ICON_VERSION)


_ICON_VERSION = "receipt-v1"

_KARGL_ICON_SVG = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="-6 -6 36 36">
  <rect x="-6" y="-6" width="36" height="36" fill="#92400e"/>
  <path d="M12 17V7" stroke="white" stroke-width="1.3" stroke-linecap="round" stroke-linejoin="round" fill="none"/>
  <path d="M16 8h-6a2 2 0 0 0 0 4h4a2 2 0 0 1 0 4H8" stroke="white" stroke-width="1.3" stroke-linecap="round" stroke-linejoin="round" fill="none"/>
  <path d="M4 3a1 1 0 0 1 1-1 1.3 1.3 0 0 1 .7.2l.933.6a1.3 1.3 0 0 0 1.4 0l.934-.6a1.3 1.3 0 0 1 1.4 0l.933.6a1.3 1.3 0 0 0 1.4 0l.933-.6a1.3 1.3 0 0 1 1.4 0l.934.6a1.3 1.3 0 0 0 1.4 0l.933-.6A1.3 1.3 0 0 1 19 2a1 1 0 0 1 1 1v18a1 1 0 0 1-1 1 1.3 1.3 0 0 1-.7-.2l-.933-.6a1.3 1.3 0 0 0-1.4 0l-.934.6a1.3 1.3 0 0 1-1.4 0l-.933-.6a1.3 1.3 0 0 0-1.4 0l-.933.6a1.3 1.3 0 0 1-1.4 0l-.934-.6a1.3 1.3 0 0 0-1.4 0l-.933.6a1.3 1.3 0 0 1-.7.2 1 1 0 0 1-1-1z" stroke="white" stroke-width="1.3" stroke-linecap="round" stroke-linejoin="round" fill="none"/>
</svg>'''


def _generate_kargl_icon(size: int, path: Path) -> None:
    import cairosvg
    cairosvg.svg2png(
        bytestring=_KARGL_ICON_SVG.encode(),
        write_to=str(path),
        output_width=size,
        output_height=size,
    )
    log(f"🖼  Kargl-Icon erstellt: {path}")


# ── Startup ───────────────────────────────────────────────────────────────────

def _startup() -> None:
    SESSIONS_DIR.mkdir(exist_ok=True)
    try:
        _ensure_kargl_icons()
    except Exception as e:
        log(f"⚠️  Icon-Generierung: {e}")


_startup()


# ── Dropbox ───────────────────────────────────────────────────────────────────

def get_dropbox_client() -> dropbox.Dropbox:
    return dropbox.Dropbox(
        oauth2_refresh_token=DROPBOX_INVOICE_REFRESH_TOKEN,
        app_key=DROPBOX_INVOICE_APP_KEY,
        app_secret=DROPBOX_INVOICE_APP_SECRET,
    )


# ── Claude OCR ────────────────────────────────────────────────────────────────

def extract_invoice_data(file_path: str, suffix: str) -> dict:
    data       = base64.standard_b64encode(Path(file_path).read_bytes()).decode("utf-8")
    media_type = MEDIA_TYPES.get(suffix, "application/pdf")

    system_prompt = (
        "Du bist ein präzises OCR- und Adressvalidierungs-System für handgeschriebene "
        "deutsche Rechnungszettel des Betriebs Josef Kargl, Holzimprägnierwerk, Traich 2, "
        "84101 Obersüßbach (Inhaber: Reinhard Kargl).\n\n"
        "WICHTIG – ABSENDER vs. EMPFÄNGER:\n"
        "Die Zettel sind auf vorgedrucktem Papier geschrieben. Oben links steht gedruckt "
        "'Kargl Reinhard' – das ist der ABSENDER des Betriebs, NICHT der Kunde. "
        "Ignoriere alle gedruckten Texte (Briefkopf, Vordrucke). "
        "Lies ausschließlich die HANDGESCHRIEBENEN Inhalte auf dem Papier. "
        "Der Kundenname und die Kundenadresse stehen immer handgeschrieben.\n\n"
        "ADRESS-VALIDIERUNG – sehr wichtig:\n"
        "Prüfe jeden Adressteil kritisch auf Plausibilität:\n"
        "- Straße: Existiert dieser Straßenname realistisch in Deutschland? "
        "Typische OCR-Fehler bei Handschrift: 'g'→'o', 'n'→'u', 'ei'→'ai', 'rn'→'m' etc. "
        "Beispiele: 'Kirchenwey'→'Kirchenweg', 'Unterolaim'→'Unterglaim', 'Mainbry'→'Mainburg'. "
        "Wenn ein Straßenname ungewöhnliche Buchstabenkombinationen enthält die kein echtes "
        "deutsches Wort ergeben: korrigieren UND address_uncertain=true setzen.\n"
        "- PLZ + Ort: Passt die PLZ zur Region des Ortes? "
        "84xxx = Niederbayern (Landshut, Dingolfing, Mainburg etc.). "
        "Ergolding (84030) liegt bei Landshut – plausibel prüfen.\n"
        "- address_uncertain=true NUR in diesen Fällen: (1) du hast einen Buchstaben korrigiert "
        "weil er als OCR-Fehler erkennbar war, (2) der Straßenname oder Ort ist unleserlich, "
        "(3) PLZ und Ort passen geografisch nicht zusammen. "
        "address_uncertain=false wenn die Adresse klar lesbar und geografisch plausibel ist – "
        "auch wenn die Handschrift etwas schwer lesbar war aber eindeutig entzifferbar.\n\n"
        "WICHTIG: Berechne KEINE Summen selbst. Lies nur die Rohdaten vom Zettel ab.\n\n"
        "Gib AUSSCHLIESSLICH ein valides JSON-Objekt zurück – keinen weiteren Text, "
        "keine Erklärungen, keine Markdown-Backticks."
    )

    user_prompt = (
        "Extrahiere alle handgeschriebenen Daten aus diesem Rechnungszettel "
        "und gib folgendes JSON zurück:\n"
        "{\n"
        "  \"anrede\": \"Firma | Herr | Frau\",\n"
        "  \"name\": \"vollständiger Name des KUNDEN (handgeschrieben, nicht der gedruckte Briefkopf)\",\n"
        "  \"strasse_nr\": \"Straße und Hausnummer des Kunden\",\n"
        "  \"plz\": \"Postleitzahl des Kunden\",\n"
        "  \"ort\": \"Ort des Kunden\",\n"
        "  \"address_uncertain\": false,\n"
        "  \"beschreibungstext\": \"Beschreibungstext für die Rechnung. "
        "Format: 'Wir [VERB] in Ihrem Auftrag [nachstehende/nachstehendes] [MATERIAL] [in KW X / am DD.MM.YYYY]:' – "
        "VERB: exakt vom Zettel lesen (z.B. 'imprägnierten', 'schälten', 'hobelten'). "
        "MATERIAL: exakter Materialname vom Zettel (z.B. 'Schnittholz', 'Bretter', 'Dachlatten 30x50 mm', 'Rundholz'). "
        "Grammatik beachten: 'nachstehendes' bei Neutrum (Rundholz, Schnittholz), 'nachstehende' bei Plural (Bretter, Dachlatten). "
        "Niemals Verb oder Material erfinden. Wenn kein Datum/KW lesbar: weglassen.\",\n"
        "  \"positionen\": [\n"
        "    {\"menge\": 1.66, \"einheit\": \"cbm\", \"einzelpreis\": 110.00, \"positions_beschreibung\": \"\"}\n"
        "  ],\n"
        "  \"netto_auf_zettel\": 182.60,\n"
        "  \"mwst_auf_zettel\": 34.69,\n"
        "  \"brutto_auf_zettel\": 217.29,\n"
        "  \"hinweis\": \"Betrag bereits bar bezahlt.\"\n"
        "}\n\n"
        "Hinweise:\n"
        "- name: das ist IMMER der handgeschriebene Kundenname – niemals 'Kargl' oder 'Reinhard'\n"
        "- anrede: 'Firma' wenn GbR, GmbH, AG o.ä., sonst 'Herr' oder 'Frau'\n"
        "- positionen: ein Eintrag pro Zeile auf dem Zettel, wenn Menge UND Einzelpreis angegeben sind. "
        "Einheit übernehmen wie auf dem Zettel (z.B. 'cbm', 'St.', 'lfm', 'fm'). "
        "positions_beschreibung: zusätzliche Spezifikationen der Position (z.B. '9m, 20-22 cm Zopf', "
        "'4m lang', '30x50 mm') – leer lassen wenn keine vorhanden. "
        "Wenn KEIN Einzelpreis angegeben ist (nur ein Gesamtbetrag): positionen = []\n"
        "- netto_auf_zettel / mwst_auf_zettel / brutto_auf_zettel: alle auf dem Zettel notierten "
        "Beträge eintragen (null wenn nicht lesbar/vorhanden)\n"
        "- beschreibungstext: Verb UND Material EXAKT aus der Handschrift lesen – "
        "niemals Verb oder Material erfinden; Grammatik beachten (nachstehendes/nachstehende)\n"
        "- hinweis: NUR explizite Zahlungshinweise vom Zettel (z.B. 'Betrag bereits bar bezahlt.', "
        "'Bereits überwiesen.') – Telefonnummern, E-Mail-Adressen, betriebliche Notizen und "
        "sonstige Vermerke gehören NICHT in den Hinweis. "
        "Leer lassen wenn kein Zahlungshinweis vorhanden.\n"
        "- beschreibungstext: Den Leistungstext formulieren. "
        "Wenn Verb und Material klar erkennbar sind: "
        "'Wir [VERB] in Ihrem Auftrag [nachstehende/nachstehendes] [MATERIAL] [Datum/KW]:'. "
        "Wenn kein klares Verb/Material vorhanden, sondern nur Notizen oder Referenzen "
        "(z.B. 'OKK als Ersatzteilspender', 'OKK TTH 4 Bagger'): "
        "diese Notizen direkt als Beschreibungstext übernehmen, exakt wie geschrieben. "
        "Telefonnummern und E-Mail-Adressen gehören NICHT in den Beschreibungstext.\n"
        "- address_uncertain=true NUR wenn ein Buchstabe korrigiert wurde oder der Ort "
        "nicht eindeutig lesbar war – NICHT bei klar lesbaren Adressen wie 'Industriestr. 2, "
        "94330 Salching'"
    )

    if suffix == ".pdf":
        file_content = {"type": "document", "source": {"type": "base64", "media_type": media_type, "data": data}}
    else:
        file_content = {"type": "image",    "source": {"type": "base64", "media_type": media_type, "data": data}}

    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    raw    = None
    for attempt in range(1, 4):
        try:
            message = client.messages.create(
                model=INVOICE_MODEL,
                max_tokens=1024,
                system=system_prompt,
                messages=[{"role": "user", "content": [file_content, {"type": "text", "text": user_prompt}]}],
            )
            raw = message.content[0].text.strip()
            break
        except anthropic.APIStatusError as e:
            if e.status_code == 529 and attempt < 3:
                wait = attempt * 15
                log(f"⏳  API überlastet (Versuch {attempt}/3) – warte {wait}s ...")
                time.sleep(wait)
            else:
                raise

    if "```" in raw:
        match = re.search(r"```(?:json)?\s*([\s\S]+?)```", raw)
        raw = match.group(1).strip() if match else raw

    if not raw:
        raise ValueError("Claude hat eine leere Antwort geliefert")

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        log(f"⚠️  Claude-Antwort (kein JSON): {raw[:300]}")
        raise


# ── Berechnung & Validierung ──────────────────────────────────────────────────

def calculate_and_validate(data: dict) -> dict:
    positionen    = [p for p in (data.get("positionen") or [])
                     if (p.get("menge") or p.get("menge_cbm")) and p.get("einzelpreis")]
    netto_zettel  = data.get("netto_auf_zettel")
    mwst_zettel   = data.get("mwst_auf_zettel")
    brutto_zettel = data.get("brutto_auf_zettel")

    if positionen:
        netto  = round(sum((p.get("menge") or p.get("menge_cbm")) * p["einzelpreis"] for p in positionen), 2)
        mwst   = round(netto * 0.19, 2)
        brutto = round(netto + mwst, 2)

        netto_ok  = netto_zettel  is None or abs(netto  - netto_zettel)  < 0.02
        brutto_ok = brutto_zettel is None or abs(brutto - brutto_zettel) < 0.02

        if not netto_ok:
            log(f"⚠️  Netto-Abweichung: berechnet {netto:.2f} € vs. Zettel {netto_zettel:.2f} €")
        if not brutto_ok:
            log(f"⚠️  Brutto-Abweichung: berechnet {brutto:.2f} € vs. Zettel {brutto_zettel:.2f} €")
    else:
        log("ℹ️  Pauschalbetrag-Modus: keine cbm×Preis-Positionen, Zettelwerte werden verwendet")
        brutto      = brutto_zettel or 0.0
        netto       = netto_zettel  or round(brutto / 1.19, 2)
        mwst        = mwst_zettel   or round(netto * 0.19, 2)
        brutto_calc = round(netto + mwst, 2)
        netto_ok    = True
        brutto_ok   = abs(brutto_calc - brutto) < 0.02
        if not brutto_ok:
            log(f"⚠️  Pauschalbetrag-Abweichung: {netto:.2f} + {mwst:.2f} = {brutto_calc:.2f} "
                f"≠ Zettel {brutto:.2f} €")

    return {"netto": netto, "mwst": mwst, "brutto": brutto,
            "netto_ok": netto_ok, "brutto_ok": brutto_ok, "pauschal": not bool(positionen)}


def fmt_eur(value: float) -> str:
    formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{formatted} €"


def fmt_cbm(value: float) -> str:
    return f"{value:.2f}".replace(".", ",") + " cbm"


# ── DOCX-Erstellung ───────────────────────────────────────────────────────────

def build_docx(data: dict, calc: dict, rechnungsnummer: str = "") -> str:
    tpl        = DocxTemplate(INVOICE_TEMPLATE)
    positionen = [p for p in (data.get("positionen") or [])
                  if (p.get("menge") or p.get("menge_cbm")) and p.get("einzelpreis")]

    pos_context = {}
    if positionen:
        for i in range(1, 7):
            if i <= len(positionen):
                p            = positionen[i - 1]
                menge        = p.get("menge") or p.get("menge_cbm")
                einheit      = p.get("einheit") or "cbm"
                pos_zusatz   = p.get("positions_beschreibung", "") or ""
                zeilen_netto = round(menge * p["einzelpreis"], 2)
                if einheit == "cbm":
                    pos_str = fmt_cbm(menge)
                elif menge == int(menge):
                    pos_str = f"{int(menge)} {einheit}"
                else:
                    pos_str = f"{menge:.2f}".replace(".", ",") + f" {einheit}"
                if pos_zusatz:
                    pos_str += f", {pos_zusatz}"
                pos_context[f"position{i}"]    = pos_str
                pos_context[f"einzelpreis{i}"] = fmt_eur(p["einzelpreis"])
                pos_context[f"gesamtpreis{i}"] = fmt_eur(zeilen_netto)
            else:
                pos_context[f"position{i}"]    = ""
                pos_context[f"einzelpreis{i}"] = ""
                pos_context[f"gesamtpreis{i}"] = ""
    else:
        pos_context["position1"]    = ""
        pos_context["einzelpreis1"] = ""
        pos_context["gesamtpreis1"] = fmt_eur(calc["netto"])
        for i in range(2, 7):
            pos_context[f"position{i}"]    = ""
            pos_context[f"einzelpreis{i}"] = ""
            pos_context[f"gesamtpreis{i}"] = ""

    context = {
        "rechnungsnummer":   rechnungsnummer,
        "anrede":            data.get("anrede", ""),
        "name":              data.get("name", ""),
        "strasse_nr":        data.get("strasse_nr", ""),
        "plz":               data.get("plz", ""),
        "ort":               data.get("ort", ""),
        "datum":             datetime.now().strftime("%d.%m.%Y"),
        "beschreibungstext": data.get("beschreibungstext", ""),
        "hinweis":           data.get("hinweis") or "",
        "netto":             fmt_eur(calc["netto"]),
        "mwst":              fmt_eur(calc["mwst"]),
        "brutto":            fmt_eur(calc["brutto"]),
        **pos_context,
    }

    tpl.render(context)
    tmp_path = tempfile.mktemp(suffix=".docx")
    tpl.save(tmp_path)
    return tmp_path


# ── Excel-Hilfsfunktionen ─────────────────────────────────────────────────────

def _download_excel(dbx: dropbox.Dropbox, dropbox_path: str):
    _, res = dbx.files_download(dropbox_path)
    tmp = tempfile.mktemp(suffix=".xlsx")
    Path(tmp).write_bytes(res.content)
    wb = openpyxl.load_workbook(tmp)
    Path(tmp).unlink(missing_ok=True)
    return wb


def _upload_excel(dbx: dropbox.Dropbox, wb, dropbox_path: str) -> None:
    tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    with open(tmp, "rb") as f:
        dbx.files_upload(f.read(), dropbox_path, mode=dropbox.files.WriteMode.overwrite, mute=True)
    Path(tmp).unlink(missing_ok=True)


def _ensure_address_excel(dbx: dropbox.Dropbox):
    try:
        return _download_excel(dbx, INVOICE_ADDRESS_FILE)
    except dropbox.exceptions.ApiError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Adressen"
        ws.append(_ADDR_HEADERS)
        header_fill = PatternFill("solid", fgColor="4472C4")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 15
        _upload_excel(dbx, wb, INVOICE_ADDRESS_FILE)
        log("📋  Adressen.xlsx neu angelegt")
        return wb


def find_in_address_excel(dbx: dropbox.Dropbox, name: str) -> dict | None:
    try:
        wb         = _ensure_address_excel(dbx)
        ws         = wb.active
        name_lower = name.strip().lower()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip().lower() == name_lower:
                return {"name": row[0], "strasse_nr": row[1], "plz": str(row[2]), "ort": row[3]}
    except Exception as e:
        log(f"⚠️  Adressliste lesen: {e}")
    return None


def add_to_address_excel(dbx: dropbox.Dropbox, name: str, strasse: str,
                         plz: str, ort: str, street_ok: bool, location_ok: bool,
                         street_corrected: bool = False) -> None:
    try:
        wb          = _ensure_address_excel(dbx)
        ws          = wb.active
        street_val  = "Korrigiert" if street_corrected else ("Ja" if street_ok else "Nein")
        ws.append([name, strasse, plz, ort,
                   street_val, "Ja" if location_ok else "Nein",
                   datetime.now().strftime("%Y-%m-%d")])
        _upload_excel(dbx, wb, INVOICE_ADDRESS_FILE)
        log(f"📋  Adresse hinzugefügt: {name}")
    except Exception as e:
        log(f"⚠️  Adresse speichern: {e}")


# ── Rechnungsnummer ───────────────────────────────────────────────────────────

def _format_invoice_nr(seq: int, year: int) -> str:
    return f"{seq:03d}00{year:02d}"


def _parse_invoice_nr(nr_str: str) -> tuple[int, int]:
    nr = str(nr_str).strip()
    if len(nr) == 7:
        return int(nr[:3]), int(nr[5:7])
    return 0, 0


def get_next_invoice_number(dbx: dropbox.Dropbox) -> str:
    current_year = int(datetime.now().strftime("%y"))
    try:
        wb      = _download_excel(dbx, INVOICE_REGISTER_FILE)
        ws      = wb.active
        last_nr = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                last_nr = str(row[0])
        if last_nr:
            seq, year = _parse_invoice_nr(last_nr)
            if year == current_year:
                return _format_invoice_nr(seq + 1, current_year)
            return _format_invoice_nr(1, current_year)
    except dropbox.exceptions.ApiError:
        pass
    except Exception as e:
        log(f"⚠️  Rechnungsnummer lesen: {e}")
    return _format_invoice_nr(17, current_year)


def _ensure_register_excel(dbx: dropbox.Dropbox):
    try:
        return _download_excel(dbx, INVOICE_REGISTER_FILE)
    except dropbox.exceptions.ApiError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rechnungen"
        ws.append(_REG_HEADERS)
        header_fill = PatternFill("solid", fgColor="375623")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for col, width in zip("ABCDEFGHI", [16, 12, 10, 22, 18, 45, 12, 12, 12]):
            ws.column_dimensions[col].width = width
        log("📊  Rechnungsregister.xlsx neu angelegt")
        return wb


def save_to_invoice_register(dbx: dropbox.Dropbox, rechnungsnummer: str,
                              data: dict, calc: dict) -> None:
    try:
        wb      = _ensure_register_excel(dbx)
        ws      = wb.active
        anrede  = data.get("anrede", "")
        name    = data.get("name", "")
        if anrede == "Firma":
            nachname, vorname = name, ""
        else:
            parts    = name.strip().rsplit(" ", 1)
            nachname = parts[1] if len(parts) == 2 else name
            vorname  = parts[0] if len(parts) == 2 else ""
        beschreibung = (data.get("beschreibungstext") or "")[:500]
        ws.append([
            rechnungsnummer, datetime.now().strftime("%d.%m.%Y"),
            anrede, nachname, vorname, beschreibung,
            round(calc["netto"], 2), round(calc["mwst"], 2), round(calc["brutto"], 2),
        ])
        _upload_excel(dbx, wb, INVOICE_REGISTER_FILE)
        log(f"📊  Register: {rechnungsnummer} – {name}")
    except Exception as e:
        log(f"⚠️  Rechnungsregister speichern: {e}")


# ── Adressvalidierung (Nominatim) ─────────────────────────────────────────────

def _nominatim_get(url: str) -> list:
    time.sleep(1)
    req = urllib.request.Request(url, headers={"User-Agent": "KarglRechnungsService/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read())
    except Exception as e:
        log(f"⚠️  Nominatim: {e}")
        return []


def validate_and_correct_address(strasse_nr: str, plz: str, ort: str) -> dict:
    query   = f"{strasse_nr}, {plz} {ort}, Deutschland"
    url     = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"q": query, "format": "json", "limit": "1", "countrycodes": "de", "addressdetails": "1"})
    results = _nominatim_get(url)

    out = {"street_ok": False, "strasse_nr": strasse_nr, "plz": plz, "ort": ort, "corrected": False}
    if not results:
        return out

    out["street_ok"] = True
    addr = results[0].get("address", {})

    osm_road = addr.get("road", "")
    osm_hnr  = addr.get("house_number", "")
    osm_plz  = addr.get("postcode", "")

    ort_candidates = [v for k, v in addr.items()
                      if k in ("city", "town", "municipality", "village", "hamlet", "suburb", "district") and v]
    if ort_candidates:
        best_ort   = max(ort_candidates, key=lambda c: difflib.SequenceMatcher(None, ort.lower(), c.lower()).ratio())
        best_ratio = difflib.SequenceMatcher(None, ort.lower(), best_ort.lower()).ratio()
        osm_ort    = best_ort if best_ratio >= 0.6 else ""
    else:
        osm_ort = ""

    m         = re.match(r"^(.*?)\s+(\d+\w*)$", strasse_nr.strip())
    input_hnr = m.group(2) if m else ""

    hnr         = osm_hnr or input_hnr
    new_strasse = f"{osm_road} {hnr}".strip() if osm_road else strasse_nr
    new_plz     = osm_plz or plz
    new_ort     = osm_ort or ort

    if (new_strasse.lower() != strasse_nr.strip().lower()
            or new_plz != plz
            or new_ort.lower() != ort.lower()):
        out["corrected"] = True

    out["strasse_nr"] = new_strasse
    out["plz"]        = new_plz
    out["ort"]        = new_ort
    return out


def validate_location_nominatim(plz: str, ort: str) -> bool:
    query = f"{plz} {ort}, Deutschland"
    url   = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"q": query, "format": "json", "limit": "1", "countrycodes": "de"})
    return bool(_nominatim_get(url))


def lookup_company_nominatim(firma: str, strasse_nr: str, plz: str, ort: str) -> str | None:
    for query in [f"{firma}, {strasse_nr}, {plz} {ort}", f"{firma}, {ort}, Deutschland"]:
        url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
            {"q": query, "format": "json", "limit": "3", "countrycodes": "de"})
        results = _nominatim_get(url)
        if results:
            osm_name = results[0].get("name", "").strip()
            if osm_name:
                ratio = difflib.SequenceMatcher(None, firma.lower(), osm_name.lower()).ratio()
                if ratio >= 0.5:
                    return osm_name
    return None


def enrich_invoice_address(dbx: dropbox.Dropbox, data: dict) -> dict:
    name  = data.get("name", "").strip()
    known = find_in_address_excel(dbx, name)
    if known:
        data["strasse_nr"]        = known["strasse_nr"]
        data["plz"]               = known["plz"]
        data["ort"]               = known["ort"]
        data["address_uncertain"] = False
        log(f"📋  Adresse aus Liste übernommen: {name}")
        return data

    strasse     = data.get("strasse_nr", "")
    plz         = data.get("plz", "")
    ort         = data.get("ort", "")
    location_ok = validate_location_nominatim(plz, ort)
    addr        = validate_and_correct_address(strasse, plz, ort)

    if not location_ok:
        data["address_uncertain"] = True
        log(f"⚠️  PLZ+Ort nicht verifiziert: {plz} {ort}")
    elif not addr["street_ok"]:
        data["address_uncertain"] = True
        log(f"⚠️  Straße nicht verifiziert: {strasse}, {plz} {ort}")
    else:
        if addr["corrected"]:
            log(f"🔧  Adresse korrigiert: '{strasse}, {plz} {ort}' → "
                f"'{addr['strasse_nr']}, {addr['plz']} {addr['ort']}'")
            data["address_uncertain"] = True
        else:
            log(f"✅  Adresse bestätigt: {addr['strasse_nr']}, {addr['plz']} {addr['ort']}")
        data["strasse_nr"] = addr["strasse_nr"]
        data["plz"]        = addr["plz"]
        data["ort"]        = addr["ort"]

    if data.get("anrede") == "Firma":
        osm_name = lookup_company_nominatim(
            name, data.get("strasse_nr", strasse), data.get("plz", plz), data.get("ort", ort))
        if osm_name and osm_name != name:
            log(f"🏢  Firmenname korrigiert: '{name}' → '{osm_name}'")
            data["name"] = osm_name
            name = osm_name
        elif not osm_name:
            data["address_uncertain"] = True
            log(f"⚠️  Firma nicht in OSM gefunden: {name}")

    add_to_address_excel(dbx, name,
                         data.get("strasse_nr", strasse), data.get("plz", plz), data.get("ort", ort),
                         addr["street_ok"], location_ok, addr["corrected"])
    return data


# ── Fehlerbehandlung ──────────────────────────────────────────────────────────

def _move_to_error(dbx: dropbox.Dropbox, dropbox_path: str, filename: str) -> None:
    try:
        error_path = f"{INVOICE_ERROR_FOLDER}/{filename}"
        dbx.files_move_v2(dropbox_path, error_path, autorename=True)
        log(f"⚠️  Nach Fehler verschoben: {INVOICE_ERROR_FOLDER}/{filename}")
    except Exception as e:
        log(f"❌  Konnte Datei nicht nach Fehler-Ordner verschieben: {e}")


# ── Hauptverarbeitung ─────────────────────────────────────────────────────────

def process_invoice(dbx: dropbox.Dropbox, dropbox_path: str) -> None:
    filename = Path(dropbox_path).name
    suffix   = Path(dropbox_path).suffix.lower()

    if suffix not in {".pdf", ".jpg", ".jpeg", ".png", ".heic", ".heif"}:
        log(f"ℹ️  Invoice: überspringe {filename} (kein unterstütztes Format)")
        return
    if suffix in {".heic", ".heif"} and not _HEIC_SUPPORTED:
        log(f"⚠️  Invoice: HEIC-Datei {filename} empfangen, aber pillow-heif nicht installiert – überspringe")
        return

    log(f"🧾  Starte Rechnungsverarbeitung: {filename}")

    tmp_file      = None
    tmp_docx      = None
    tmp_odt       = None
    tmp_converted = None

    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f:
            tmp_file = f.name
        dbx.files_download_to_file(tmp_file, dropbox_path)
        log(f"⬇️  Heruntergeladen: {filename}")

        process_path   = tmp_file
        process_suffix = suffix
        if suffix in {".heic", ".heif"}:
            tmp_converted = tempfile.mktemp(suffix=".jpg")
            with Image.open(tmp_file) as img:
                img.convert("RGB").save(tmp_converted, format="JPEG", quality=92)
            process_path   = tmp_converted
            process_suffix = ".jpg"
            log(f"🔄  HEIC → JPEG konvertiert: {filename}")

        invoice_data    = extract_invoice_data(process_path, process_suffix)
        n_pos           = len(invoice_data.get("positionen") or [])
        log(f"🤖  Extrahiert: {invoice_data.get('name', '?')} – {n_pos} Position(en)")

        rechnungsnummer = get_next_invoice_number(dbx)
        log(f"🔢  Rechnungsnummer: {rechnungsnummer}")

        invoice_data = enrich_invoice_address(dbx, invoice_data)
        calc         = calculate_and_validate(invoice_data)
        tmp_docx     = build_docx(invoice_data, calc, rechnungsnummer)
        tmp_odt, out_ext = convert_to_odt(tmp_docx)

        clean_name   = re.sub(r"[^\w\-]", "_", invoice_data.get("name", "Unbekannt"))
        datum_str    = datetime.now().strftime("%Y-%m-%d")
        brutto_str   = f"{calc['brutto']:.2f}"
        needs_prufen = (invoice_data.get("address_uncertain")
                        or not calc["netto_ok"] or not calc["brutto_ok"])
        nr_prefix    = f"_prüfen_{rechnungsnummer}" if needs_prufen else rechnungsnummer
        out_name     = f"{nr_prefix}_Rechnung_{clean_name}_{datum_str}_{brutto_str}€{out_ext}"
        out_path     = f"{INVOICE_OUTPUT_FOLDER}/{out_name}"

        with open(tmp_odt, "rb") as fh:
            dbx.files_upload(fh.read(), out_path, mode=dropbox.files.WriteMode.overwrite, mute=True)
        log(f"⬆️  Hochgeladen: {out_path}")

        save_to_invoice_register(dbx, rechnungsnummer, invoice_data, calc)

        done_name = f"{rechnungsnummer}_Eingang_{clean_name}_{datum_str}_{brutto_str}€{suffix}"
        done_path = f"{INVOICE_DONE_FOLDER}/{done_name}"
        dbx.files_move_v2(dropbox_path, done_path, autorename=True)
        log(f"📁  Archiviert: {done_path}")

    except json.JSONDecodeError as e:
        log(f"❌  JSON-Fehler bei {filename}: {e}")
        _move_to_error(dbx, dropbox_path, filename)
    except Exception as e:
        log(f"❌  Fehler bei Rechnungsverarbeitung {filename}: {e}")
        _move_to_error(dbx, dropbox_path, filename)
    finally:
        if tmp_file:      Path(tmp_file).unlink(missing_ok=True)
        if tmp_converted: Path(tmp_converted).unlink(missing_ok=True)
        if tmp_docx:      Path(tmp_docx).unlink(missing_ok=True)
        if tmp_odt and tmp_odt != tmp_docx:
            Path(tmp_odt).unlink(missing_ok=True)


# ── Dropbox-Cursor ────────────────────────────────────────────────────────────

def get_invoice_cursor(dbx: dropbox.Dropbox) -> str:
    if Path(INVOICE_CURSOR_FILE).exists():
        return Path(INVOICE_CURSOR_FILE).read_text().strip()
    result = dbx.files_list_folder(INVOICE_INPUT_FOLDER, recursive=False)
    cursor = result.cursor
    while result.has_more:
        result = dbx.files_list_folder_continue(cursor)
        cursor = result.cursor
    Path(INVOICE_CURSOR_FILE).write_text(cursor)
    log("🔖  Invoice-Cursor gespeichert")
    return cursor


def save_invoice_cursor(cursor: str) -> None:
    Path(INVOICE_CURSOR_FILE).write_text(cursor)


def process_invoice_changes(dbx: dropbox.Dropbox) -> None:
    cursor = get_invoice_cursor(dbx)
    result = dbx.files_list_folder_continue(cursor)

    while True:
        for entry in result.entries:
            if isinstance(entry, (DeletedMetadata, FolderMetadata)):
                continue
            if str(Path(entry.path_lower).parent) != INVOICE_INPUT_FOLDER.lower():
                continue
            try:
                process_invoice(dbx, entry.path_display)
            except Exception as e:
                log(f"❌  Invoice-Fehler bei {entry.name}: {e}")

        save_invoice_cursor(result.cursor)
        if not result.has_more:
            break
        result = dbx.files_list_folder_continue(result.cursor)


# ── Flask-Routen ──────────────────────────────────────────────────────────────

@app.route("/webhook-invoice", methods=["GET"])
def verify_invoice():
    challenge = request.args.get("challenge")
    if challenge:
        return challenge, 200, {"Content-Type": "text/plain"}
    abort(400)


@app.route("/webhook-invoice", methods=["POST"])
def webhook_invoice():
    signature = request.headers.get("X-Dropbox-Signature", "")
    expected  = hmac.new(
        DROPBOX_INVOICE_APP_SECRET.encode(),
        request.data,
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(signature, expected):
        log("⚠️  Ungültige Invoice-Webhook-Signatur")
        abort(403)

    dbx = get_dropbox_client()
    threading.Thread(target=process_invoice_changes, args=(dbx,), daemon=True).start()
    return "", 200


# ── Kargl App ─────────────────────────────────────────────────────────────────

@app.route("/kargl/")
@app.route("/kargl")
def kargl_index():
    if KARGL_HTML_PATH.exists():
        return KARGL_HTML_PATH.read_text("utf-8"), 200, {"Content-Type": "text/html; charset=utf-8"}
    return "App nicht gefunden", 404


@app.route("/kargl/manifest.json")
def kargl_manifest():
    return {
        "name": "Kargl Rechnung",
        "short_name": "Kargl",
        "start_url": "/kargl/",
        "display": "standalone",
        "background_color": "#1c1c1e",
        "theme_color": "#92400e",
        "icons": [
            {"src": "/kargl/icon-192.png?v=3", "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": "/kargl/icon-512.png?v=3", "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": "/kargl/icon-180.png?v=3", "sizes": "180x180", "type": "image/png"},
        ],
    }


@app.route("/kargl/sw.js")
def kargl_sw():
    sw = (
        "const CACHE='kargl-v4';\n"
        "const SHELL=['/kargl/','/kargl/manifest.json','/kargl/icon-192.png?v=3','/kargl/icon-512.png?v=3','/kargl/icon-180.png?v=3'];\n"
        "self.addEventListener('install',e=>{e.waitUntil(caches.open(CACHE).then(c=>c.addAll(SHELL)));self.skipWaiting();});\n"
        "self.addEventListener('activate',e=>{e.waitUntil(caches.keys().then(keys=>Promise.all(keys.filter(k=>k!==CACHE).map(k=>caches.delete(k)))));self.clients.claim();});\n"
        "self.addEventListener('fetch',e=>{\n"
        "  const url=new URL(e.request.url);\n"
        "  if(url.pathname.startsWith('/kargl/api/'))return;\n"
        "  if(e.request.destination==='document'){e.respondWith(fetch(e.request).catch(()=>caches.match('/kargl/')));return;}\n"
        "  e.respondWith(caches.match(e.request).then(c=>c||fetch(e.request)));\n"
        "});\n"
    )
    return sw, 200, {
        "Content-Type": "application/javascript",
        "Cache-Control": "no-cache, no-store",
        "Service-Worker-Allowed": "/kargl/",
    }


@app.route("/kargl/icon-<size>.png")
def kargl_icon(size):
    p = KARGL_ICONS_DIR / f"icon-{size}.png"
    if not p.exists():
        abort(404)
    return p.read_bytes(), 200, {"Content-Type": "image/png", "Cache-Control": "public, max-age=86400"}


@app.route("/kargl/api/auth", methods=["POST"])
def kargl_auth():
    body  = request.json or {}
    token = body.get("token", "")
    if not KARGL_APP_TOKEN or token != KARGL_APP_TOKEN:
        return {"error": "Ungültiger Code"}, 401
    return {"ok": True}


@app.route("/kargl/api/ocr", methods=["POST"])
@require_token
def kargl_ocr():
    f = request.files.get("file")
    if not f:
        return {"error": "Kein Bild"}, 400

    suffix = Path(f.filename or "x.jpg").suffix.lower()
    if suffix not in {".pdf", ".jpg", ".jpeg", ".png", ".heic", ".heif"}:
        return {"error": "Nicht unterstütztes Format"}, 400

    SESSIONS_DIR.mkdir(exist_ok=True)
    _cleanup_old_sessions()

    session_id = uuid_module.uuid4().hex
    img_path   = SESSIONS_DIR / f"{session_id}{suffix}"
    f.save(str(img_path))

    process_path   = str(img_path)
    process_suffix = suffix
    if suffix in {".heic", ".heif"} and _HEIC_SUPPORTED:
        jpg_path = str(SESSIONS_DIR / f"{session_id}.jpg")
        with Image.open(process_path) as img:
            img.convert("RGB").save(jpg_path, format="JPEG", quality=92)
        img_path.unlink(missing_ok=True)
        img_path       = SESSIONS_DIR / f"{session_id}.jpg"
        process_path   = str(img_path)
        process_suffix = ".jpg"
        suffix         = ".jpg"

    try:
        invoice_data = extract_invoice_data(process_path, process_suffix)
        log(f"🤖  App OCR: {invoice_data.get('name', '?')}")

        dbx          = get_dropbox_client()
        invoice_data = enrich_invoice_address(dbx, invoice_data)
        log(f"📋  App Adresse: {invoice_data.get('strasse_nr', '?')}, {invoice_data.get('plz', '?')} {invoice_data.get('ort', '?')}")

        meta = {"original_suffix": suffix, "created_at": datetime.now().isoformat()}
        (SESSIONS_DIR / f"{session_id}.json").write_text(json.dumps(meta, ensure_ascii=False))

        next_nr = get_next_invoice_number(dbx)
        return {"session_id": session_id, "fields": invoice_data, "next_rechnungsnummer": next_nr}

    except Exception as e:
        img_path.unlink(missing_ok=True)
        log(f"❌  App OCR Fehler: {e}")
        return {"error": f"OCR fehlgeschlagen: {e}"}, 500


@app.route("/kargl/api/confirm", methods=["POST"])
@require_token
def kargl_confirm():
    body       = request.json or {}
    session_id = body.get("session_id", "")
    fields     = body.get("fields", {})
    custom_nr  = (body.get("rechnungsnummer") or "").strip()

    if not session_id:
        return {"error": "Keine Session"}, 400

    session_file = SESSIONS_DIR / f"{session_id}.json"
    if not session_file.exists():
        return {"error": "Session abgelaufen – bitte neu scannen"}, 404

    meta       = json.loads(session_file.read_text())
    img_suffix = meta.get("original_suffix", ".jpg")
    img_path   = SESSIONS_DIR / f"{session_id}{img_suffix}"

    dbx  = get_dropbox_client()
    calc = calculate_and_validate(fields)
    name = fields.get("name", "")

    rechnungsnummer = custom_nr if custom_nr else get_next_invoice_number(dbx)

    tmp_docx = None
    tmp_odt  = None
    tmp_pdf  = None
    try:
        tmp_docx = build_docx(fields, calc, rechnungsnummer)
        tmp_odt, out_ext = convert_to_odt(tmp_docx)

        clean_name   = re.sub(r"[^\w\-]", "_", name or "Unbekannt")
        datum_str    = datetime.now().strftime("%Y-%m-%d")
        brutto_str   = f"{calc['brutto']:.2f}"
        needs_prufen = (
            fields.get("address_uncertain")
            or not calc["netto_ok"]
            or not calc["brutto_ok"]
        )
        nr_prefix = f"_prüfen_{rechnungsnummer}" if needs_prufen else rechnungsnummer
        out_name  = f"{nr_prefix}_Rechnung_{clean_name}_{datum_str}_{brutto_str}€{out_ext}"
        out_path  = f"{INVOICE_OUTPUT_FOLDER}/{out_name}"

        with open(tmp_odt, "rb") as fh:
            dbx.files_upload(fh.read(), out_path, mode=dropbox.files.WriteMode.overwrite, mute=True)
        log(f"⬆️  App: Hochgeladen: {out_path}")

        save_to_invoice_register(dbx, rechnungsnummer, fields, calc)

        # Original-Bild archivieren
        if img_path.exists():
            done_name = f"{rechnungsnummer}_Eingang_{clean_name}_{datum_str}_{brutto_str}€{img_suffix}"
            done_path = f"{INVOICE_DONE_FOLDER}/{done_name}"
            with open(img_path, "rb") as fh:
                dbx.files_upload(fh.read(), done_path, mode=dropbox.files.WriteMode.overwrite, mute=True)
            log(f"📁  App: Original archiviert: {done_path}")

        # PDF für In-App-Viewer generieren
        pdf_session_id = None
        tmp_pdf = convert_to_pdf(tmp_odt)
        if tmp_pdf:
            pdf_dest = SESSIONS_DIR / f"{session_id}.pdf"
            shutil.move(tmp_pdf, str(pdf_dest))
            tmp_pdf = None  # bereits verschoben
            pdf_session_id = session_id
            log(f"📄  PDF erstellt: {pdf_dest.name}")

        session_file.unlink(missing_ok=True)
        img_path.unlink(missing_ok=True)

        log(f"✅  App: Rechnung erstellt: {out_name}")
        return {"out_name": out_name, "rechnungsnummer": rechnungsnummer,
                "pdf_session_id": pdf_session_id}

    except Exception as e:
        log(f"❌  App Confirm Fehler: {e}")
        return {"error": str(e)}, 500
    finally:
        if tmp_docx:                          Path(tmp_docx).unlink(missing_ok=True)
        if tmp_odt and tmp_odt != tmp_docx:   Path(tmp_odt).unlink(missing_ok=True)
        if tmp_pdf:                           Path(tmp_pdf).unlink(missing_ok=True)


# ── PDF-Session-Endpoint ──────────────────────────────────────────────────────

@app.route("/kargl/api/sessions/<sid>/pdf", methods=["GET"])
@require_token_or_param
def kargl_session_pdf(sid):
    if not re.match(r'^[a-f0-9]{32}$', sid):
        abort(400)
    pdf_path = SESSIONS_DIR / f"{sid}.pdf"
    if not pdf_path.exists():
        return {"error": "PDF nicht gefunden oder abgelaufen"}, 404
    pdf_data = pdf_path.read_bytes()
    return Response(pdf_data, mimetype="application/pdf",
                    headers={"Content-Disposition": 'inline; filename="rechnung.pdf"',
                             "Cache-Control": "no-store"})


# ── Adressen-Endpoints ────────────────────────────────────────────────────────

@app.route("/kargl/api/adressen", methods=["GET"])
@require_token
def kargl_adressen_list():
    dbx = get_dropbox_client()
    try:
        wb = _ensure_address_excel(dbx)
        ws = wb.active
        adressen = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if any(cell is not None and str(cell).strip() for cell in row[:4]):
                adressen.append({
                    "row": i,
                    "name":    str(row[0] or "").strip(),
                    "strasse": str(row[1] or "").strip(),
                    "plz":     str(row[2] or "").strip(),
                    "ort":     str(row[3] or "").strip(),
                })
        return {"adressen": adressen}
    except Exception as e:
        log(f"⚠️  Adressen lesen: {e}")
        return {"error": str(e)}, 500


@app.route("/kargl/api/adressen/<int:row>", methods=["POST"])
@require_token
def kargl_adressen_update(row):
    body = request.json or {}
    dbx  = get_dropbox_client()
    try:
        wb = _ensure_address_excel(dbx)
        ws = wb.active
        ws.cell(row=row, column=1).value = body.get("name", "")
        ws.cell(row=row, column=2).value = body.get("strasse", "")
        ws.cell(row=row, column=3).value = body.get("plz", "")
        ws.cell(row=row, column=4).value = body.get("ort", "")
        _upload_excel(dbx, wb, INVOICE_ADDRESS_FILE)
        log(f"📋  Adresse aktualisiert: Zeile {row}")
        return {"ok": True}
    except Exception as e:
        log(f"⚠️  Adresse aktualisieren: {e}")
        return {"error": str(e)}, 500


@app.route("/kargl/api/adressen/<int:row>/loeschen", methods=["POST"])
@require_token
def kargl_adressen_delete(row):
    dbx = get_dropbox_client()
    try:
        wb = _ensure_address_excel(dbx)
        ws = wb.active
        ws.delete_rows(row)
        _upload_excel(dbx, wb, INVOICE_ADDRESS_FILE)
        log(f"📋  Adresse gelöscht: Zeile {row}")
        return {"ok": True}
    except Exception as e:
        log(f"⚠️  Adresse löschen: {e}")
        return {"error": str(e)}, 500


# ── Rechnungsordner-Endpoints ─────────────────────────────────────────────────

def _list_folder_names(dbx: dropbox.Dropbox, path: str) -> set[str]:
    """Listet alle Dateinamen in einem Dropbox-Ordner (flach)."""
    names: set[str] = set()
    try:
        result = dbx.files_list_folder(path)
        while True:
            for e in result.entries:
                names.add(e.name)
            if not result.has_more:
                break
            result = dbx.files_list_folder_continue(result.cursor)
    except Exception as e:
        log(f"⚠️  Ordner auflisten {path}: {e}")
    return names


@app.route("/kargl/api/rechnungen", methods=["GET"])
@require_token
def kargl_rechnungen_list():
    dbx = get_dropbox_client()
    try:
        wb = _ensure_register_excel(dbx)
        ws = wb.active
        rechnungen = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                rechnungen.append({
                    "nr":       str(row[0] or ""),
                    "datum":    str(row[1] or ""),
                    "anrede":   str(row[2] or ""),
                    "nachname": str(row[3] or ""),
                    "vorname":  str(row[4] or ""),
                    "brutto":   float(row[8]) if row[8] else 0,
                    "status":   "unknown",
                })
        rechnungen.reverse()
        rechnungen = rechnungen[:60]

        # Datei-Status: einmal beide Ordner listen, dann abgleichen
        entwurf_files  = _list_folder_names(dbx, INVOICE_OUTPUT_FOLDER)
        erledigt_files = _list_folder_names(dbx, INVOICE_DONE_FOLDER)

        def _get_status(nr: str) -> str:
            for name in entwurf_files:
                if nr in name and "rechnung" in name.lower() and \
                        (name.lower().endswith(".odt") or name.lower().endswith(".docx")):
                    return "entwurf"
            for name in erledigt_files:
                if nr in name and "rechnung" in name.lower() and \
                        (name.lower().endswith(".odt") or name.lower().endswith(".docx")):
                    return "erledigt"
            return "not_found"

        for r in rechnungen:
            r["status"] = _get_status(r["nr"])

        return {"rechnungen": rechnungen}
    except Exception as e:
        log(f"⚠️  Rechnungen lesen: {e}")
        return {"error": str(e)}, 500


def _find_rechnung_odt(dbx: dropbox.Dropbox, nr: str) -> str | None:
    """Sucht ODT/DOCX-Datei für Rechnungsnummer in Entwurf und Erledigt."""
    for folder in [INVOICE_OUTPUT_FOLDER, INVOICE_DONE_FOLDER]:
        try:
            result = dbx.files_list_folder(folder)
            while True:
                for entry in result.entries:
                    name = entry.name
                    if (nr in name and "rechnung" in name.lower()
                            and (name.lower().endswith(".odt") or name.lower().endswith(".docx"))):
                        return entry.path_display
                if not result.has_more:
                    break
                result = dbx.files_list_folder_continue(result.cursor)
        except Exception as e:
            log(f"⚠️  ODT suchen in {folder}: {e}")
    return None


@app.route("/kargl/api/rechnungen/<nr>/pdf", methods=["GET"])
@require_token_or_param
def kargl_rechnung_pdf(nr):
    if not re.match(r'^[\w\-_]+$', nr):
        abort(400)
    dbx      = get_dropbox_client()
    odt_path = _find_rechnung_odt(dbx, nr)
    if not odt_path:
        return {"error": "Rechnung nicht in Entwurf-Ordner gefunden"}, 404

    tmp_odt = None
    tmp_pdf = None
    try:
        tmp_odt = tempfile.mktemp(suffix=".odt")
        dbx.files_download_to_file(tmp_odt, odt_path)
        tmp_pdf = convert_to_pdf(tmp_odt)
        if not tmp_pdf:
            return {"error": "PDF-Konvertierung fehlgeschlagen"}, 500
        pdf_data = Path(tmp_pdf).read_bytes()
        return Response(pdf_data, mimetype="application/pdf",
                        headers={"Content-Disposition": f'inline; filename="{nr}_Rechnung.pdf"',
                                 "Cache-Control": "no-store"})
    except Exception as e:
        log(f"⚠️  Rechnung PDF {nr}: {e}")
        return {"error": str(e)}, 500
    finally:
        if tmp_odt: Path(tmp_odt).unlink(missing_ok=True)
        if tmp_pdf: Path(tmp_pdf).unlink(missing_ok=True)


@app.route("/kargl/api/rechnungen/<nr>/verschieben", methods=["POST"])
@require_token
def kargl_rechnung_verschieben(nr):
    if not re.match(r'^[\w\-_]+$', nr):
        abort(400)
    dbx      = get_dropbox_client()
    odt_path = _find_rechnung_odt(dbx, nr)
    if not odt_path:
        return {"error": "Rechnung nicht in Entwurf-Ordner gefunden"}, 404
    try:
        filename = Path(odt_path).name
        new_path = f"{INVOICE_DONE_FOLDER}/{filename}"
        dbx.files_move_v2(odt_path, new_path, autorename=True)
        log(f"📁  Rechnung verschoben → Erledigt: {filename}")
        return {"ok": True}
    except Exception as e:
        log(f"⚠️  Rechnung verschieben {nr}: {e}")
        return {"error": str(e)}, 500


# ── Rechnung bearbeiten ───────────────────────────────────────────────────────

@app.route("/kargl/api/rechnungen/<nr>/felder", methods=["GET"])
@require_token
def kargl_rechnung_felder(nr):
    """Gibt bekannte Felder für eine Rechnungsnummer zurück (für Bearbeiten-Flow)."""
    if not re.match(r'^[\w\-_]+$', nr):
        abort(400)
    dbx = get_dropbox_client()
    try:
        wb = _ensure_register_excel(dbx)
        ws = wb.active
        found = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0] or '').strip() == nr:
                found = row
                break
        if not found:
            return {"error": "Rechnungsnummer nicht im Register gefunden"}, 404

        anrede   = str(found[2] or 'Firma')
        nachname = str(found[3] or '')
        vorname  = str(found[4] or '')
        name     = f"{vorname} {nachname}".strip() if vorname else nachname
        beschr   = str(found[5] or '')
        brutto   = float(found[8]) if found[8] else 0.0

        adresse = find_in_address_excel(dbx, name) or {}

        fields = {
            "anrede":            anrede,
            "name":              name,
            "strasse_nr":        adresse.get("strasse_nr", ""),
            "plz":               adresse.get("plz", ""),
            "ort":               adresse.get("ort", ""),
            "beschreibungstext": beschr,
            "brutto_auf_zettel": brutto,
            "positionen":        [],
            "hinweis":           "",
            "address_uncertain": False,
        }
        return {"fields": fields, "rechnungsnummer": nr}
    except Exception as e:
        log(f"⚠️  Felder lesen {nr}: {e}")
        return {"error": str(e)}, 500


def _update_invoice_register(dbx: dropbox.Dropbox, nr: str, data: dict, calc: dict) -> None:
    try:
        wb      = _ensure_register_excel(dbx)
        ws      = wb.active
        anrede  = data.get("anrede", "")
        name    = data.get("name", "")
        if anrede == "Firma":
            nachname, vorname = name, ""
        else:
            parts    = name.strip().rsplit(" ", 1)
            nachname = parts[1] if len(parts) == 2 else name
            vorname  = parts[0] if len(parts) == 2 else ""
        beschreibung = (data.get("beschreibungstext") or "")[:500]

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value or '').strip() == nr:
                row[1].value = datetime.now().strftime("%d.%m.%Y")
                row[2].value = anrede
                row[3].value = nachname
                row[4].value = vorname
                row[5].value = beschreibung
                row[6].value = round(calc["netto"], 2)
                row[7].value = round(calc["mwst"], 2)
                row[8].value = round(calc["brutto"], 2)
                _upload_excel(dbx, wb, INVOICE_REGISTER_FILE)
                log(f"📊  Register aktualisiert: {nr}")
                return
        # Fallback: neuer Eintrag
        save_to_invoice_register(dbx, nr, data, calc)
    except Exception as e:
        log(f"⚠️  Register aktualisieren {nr}: {e}")


@app.route("/kargl/api/rechnungen/<nr>/neu-erstellen", methods=["POST"])
@require_token
def kargl_rechnung_neu_erstellen(nr):
    """Überschreibt eine bestehende Entwurf-Rechnung mit aktualisierten Feldern."""
    if not re.match(r'^[\w\-_]+$', nr):
        abort(400)
    body   = request.json or {}
    fields = body.get("fields", {})
    name   = fields.get("name", "")

    dbx  = get_dropbox_client()
    calc = calculate_and_validate(fields)

    old_path = _find_rechnung_odt(dbx, nr)

    tmp_docx = None
    tmp_odt  = None
    tmp_pdf  = None
    try:
        tmp_docx = build_docx(fields, calc, nr)
        tmp_odt, out_ext = convert_to_odt(tmp_docx)

        clean_name   = re.sub(r"[^\w\-]", "_", name or "Unbekannt")
        datum_str    = datetime.now().strftime("%Y-%m-%d")
        brutto_str   = f"{calc['brutto']:.2f}"
        needs_prufen = fields.get("address_uncertain") or not calc["netto_ok"] or not calc["brutto_ok"]
        nr_prefix    = f"_prüfen_{nr}" if needs_prufen else nr
        out_name     = f"{nr_prefix}_Rechnung_{clean_name}_{datum_str}_{brutto_str}€{out_ext}"
        out_path     = f"{INVOICE_OUTPUT_FOLDER}/{out_name}"

        # Alte Datei entfernen wenn anderer Name
        if old_path and old_path != out_path:
            try:
                dbx.files_delete_v2(old_path)
                log(f"🗑  Alte Version gelöscht: {Path(old_path).name}")
            except Exception as e:
                log(f"⚠️  Alte Datei löschen: {e}")

        with open(tmp_odt, "rb") as fh:
            dbx.files_upload(fh.read(), out_path, mode=dropbox.files.WriteMode.overwrite, mute=True)
        log(f"⬆️  Bearbeitet: {out_name}")

        _update_invoice_register(dbx, nr, fields, calc)

        # PDF für In-App-Viewer
        pdf_session_id = None
        tmp_pdf = convert_to_pdf(tmp_odt)
        if tmp_pdf:
            SESSIONS_DIR.mkdir(exist_ok=True)
            edit_sid  = uuid_module.uuid4().hex
            pdf_dest  = SESSIONS_DIR / f"{edit_sid}.pdf"
            shutil.move(tmp_pdf, str(pdf_dest))
            tmp_pdf        = None
            pdf_session_id = edit_sid

        return {"out_name": out_name, "rechnungsnummer": nr, "pdf_session_id": pdf_session_id}

    except Exception as e:
        log(f"❌  Neu-Erstellen {nr}: {e}")
        return {"error": str(e)}, 500
    finally:
        if tmp_docx:                          Path(tmp_docx).unlink(missing_ok=True)
        if tmp_odt and tmp_odt != tmp_docx:   Path(tmp_odt).unlink(missing_ok=True)
        if tmp_pdf:                           Path(tmp_pdf).unlink(missing_ok=True)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5002)
